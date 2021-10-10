# --coding:utf-8--
'''
@File: ExcelTools.py
@Author:魏林栋（welindong）
@Time: Y E A R 年 {YEAR}年YEAR年{MONTH}月D A Y 日 {DAY}日DAY日{HOUR}
'''
import openpyxl


class ExcelTools(object):
    def __init__(self, path, time):
        self.address = path + '/record.xlsx'
        self.nowtime = time
        self.excel = openpyxl.load_workbook(self.address)
        self.sheetnames = self.excel.sheetnames
        self.table = self.excel[self.sheetnames[0]]
        # self.addSheet()


    def writeInfo(self, time, word, transform, renum, wrnum):
        '''
        :param sheet: 页面名称，及日期
        :param word: 单词
        :param transform: 翻译
        :param renum: 听写次数
        :param wrnum: 错误次数
        '''
        nrows = self.table.max_row
        if self.table.cell(1, 1).value is None:
            self.write(1, word, transform, time, renum, wrnum)
            self.excel.save(self.address)
            return
        self.write(nrows + 1, word, transform, time, renum, wrnum)
        self.excel.save(self.address)


    def write(self, row, word, transform, time, renum, wrnum):
        self.table.cell(row, 1).value = word
        self.table.cell(row, 2).value = transform
        self.table.cell(row, 3).value = time
        self.table.cell(row, 4).value = renum
        self.table.cell(row, 5).value = wrnum


    # def addSheet(self):
    #     if self.sheetnames[-1] != self.nowtime:
    #         self.excel.create_sheet(self.nowtime)
    #         self.excel.save(self.address)


    def readExcel(self):
        allwords = []
        recordict = {}
        for row in range(1, self.table.max_row + 1):
            if self.table.cell(row, 1).value is not None:
                allwords.append(self.table.cell(row, 1).value + ',' + self.table.cell(row, 2).value)
                recordict[self.table.cell(row, 1).value] = [row, self.table.cell(row, 2).value, self.table.cell(row, 3).value, self.table.cell(row, 4).value, self.table.cell(row, 5).value]
        return allwords, recordict


    def record(self, row, column, info):
        self.table.cell(row, column).value = info
        self.excel.save(self.address)